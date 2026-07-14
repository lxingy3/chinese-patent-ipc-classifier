from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


IPC_CODE_RE = re.compile(r"^[A-HY]\d{2}[A-Z]\d+/\d+$")
EXTRA_FIELDS = [
    "publication_numbers",
    "ipc_codes",
    "ipc_sections",
    "ipc_classes",
    "ipc_subclasses",
    "ipc_code_count",
]


def normalize_text(value: object) -> str:
    return re.sub(r"\s+", "", str(value or ""))


def normalize_codes(values: Iterable[object]) -> list[str]:
    codes: list[str] = []
    for value in values:
        code = re.sub(r"\s+", "", str(value or "").upper())
        if not code:
            continue
        if not IPC_CODE_RE.fullmatch(code):
            raise ValueError(f"Invalid IPC code: {value!r}")
        if code not in codes:
            codes.append(code)
    return codes


def load_lookup_results(results_dir: Path) -> dict[int, dict[str, Any]]:
    candidates: dict[int, list[dict[str, Any]]] = defaultdict(list)
    paths = sorted(results_dir.glob("cnipa_part_*.json"))
    if not paths:
        raise ValueError(f"No cnipa_part_*.json files found in {results_dir}")

    for path in paths:
        records = json.loads(path.read_text(encoding="utf-8"))
        for record in records:
            candidates[int(record["row"])].append(record)

    selected: dict[int, dict[str, Any]] = {}
    for row, records in candidates.items():
        successful_sets = {
            tuple(sorted(normalize_codes(record.get("ipc_codes") or [])))
            for record in records
            if record.get("status") == "ok"
        }
        successful_sets.discard(())
        if len(successful_sets) > 1:
            raise ValueError(f"Conflicting successful IPC results for workbook row {row}")

        selected[row] = max(
            records,
            key=lambda record: (
                record.get("status") == "ok",
                len(normalize_codes(record.get("ipc_codes") or [])),
                bool(record.get("found_publication")),
            ),
        )
    return selected


def load_workbook_rows(workbook_path: Path) -> dict[tuple[str, str], list[dict[str, Any]]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook.active
    headers = [str(value or "") for value in next(sheet.iter_rows(values_only=True))]
    required = ["专利名称", "专利摘要", "授权号"]
    missing = [field for field in required if field not in headers]
    if missing:
        raise ValueError(f"Workbook is missing columns: {', '.join(missing)}")

    indices = {field: headers.index(field) for field in required}
    rows: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        title = str(values[indices["专利名称"]] or "").strip()
        abstract = str(values[indices["专利摘要"]] or "").strip()
        rows[(normalize_text(title), normalize_text(abstract))].append(
            {
                "row": row_number,
                "publication": str(values[indices["授权号"]] or "").strip(),
            }
        )
    return rows


def read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open(encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file)
        if reader.fieldnames is None:
            raise ValueError(f"CSV has no header: {path}")
        return list(reader.fieldnames), list(reader)


def ordered_prefixes(codes: Iterable[str], length: int) -> list[str]:
    values: list[str] = []
    for code in codes:
        value = code[:length]
        if value not in values:
            values.append(value)
    return values


def build_rows(
    clean_rows: list[dict[str, str]],
    workbook_rows: dict[tuple[str, str], list[dict[str, Any]]],
    lookup_results: dict[int, dict[str, Any]],
) -> list[dict[str, str]]:
    output: list[dict[str, str]] = []
    for clean_row in clean_rows:
        key = (normalize_text(clean_row["patent_title"]), normalize_text(clean_row["patent_abstract"]))
        source_rows = workbook_rows.get(key)
        if not source_rows:
            raise ValueError(f"No source workbook row for patent: {clean_row['patent_title']!r}")

        code_sets: list[list[str]] = []
        publications: list[str] = []
        for source_row in source_rows:
            result = lookup_results.get(source_row["row"])
            if result is None:
                raise ValueError(f"No IPC lookup result for workbook row {source_row['row']}")
            codes = normalize_codes(result.get("ipc_codes") or [])
            if codes:
                code_sets.append(codes)
            publication = source_row["publication"]
            if publication and publication not in publications:
                publications.append(publication)

        if not code_sets:
            raise ValueError(f"No IPC codes for patent: {clean_row['patent_title']!r}")
        canonical_sets = {tuple(sorted(codes)) for codes in code_sets}
        if len(canonical_sets) != 1:
            raise ValueError(f"Duplicate patent rows disagree on IPC codes: {clean_row['patent_title']!r}")

        codes = code_sets[0]
        primary_code = re.sub(r"\s+", "", clean_row["ipc_level_4"].upper())
        if primary_code not in codes:
            raise ValueError(f"Primary IPC code {primary_code!r} is absent from full code list")

        enriched = dict(clean_row)
        enriched.update(
            {
                "publication_numbers": "|".join(publications),
                "ipc_codes": "|".join(codes),
                "ipc_sections": "|".join(ordered_prefixes(codes, 1)),
                "ipc_classes": "|".join(ordered_prefixes(codes, 3)),
                "ipc_subclasses": "|".join(ordered_prefixes(codes, 4)),
                "ipc_code_count": str(len(codes)),
            }
        )
        output.append(enriched)
    return output


def dataset_stats(rows: list[dict[str, str]]) -> dict[str, Any]:
    code_counts = [int(row["ipc_code_count"]) for row in rows]
    level_fields = {
        "sections": "ipc_sections",
        "classes": "ipc_classes",
        "subclasses": "ipc_subclasses",
    }
    levels: dict[str, Any] = {}
    for level, field in level_fields.items():
        support = Counter(label for row in rows for label in row[field].split("|") if label)
        levels[level] = {
            "labels": len(support),
            "labels_with_at_least_5_documents": sum(count >= 5 for count in support.values()),
        }
    return {
        "documents": len(rows),
        "multilabel_documents": sum(count > 1 for count in code_counts),
        "mean_ipc_codes_per_document": sum(code_counts) / len(code_counts),
        "max_ipc_codes_per_document": max(code_counts),
        "levels": levels,
    }


def write_csv(path: Path, fields: list[str], rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    parser = argparse.ArgumentParser(description="Attach complete IPC assignments to the cleaned patent corpus.")
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument("--results-dir", type=Path, required=True)
    parser.add_argument("--clean-data", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--stats-output", type=Path)
    args = parser.parse_args()

    fields, clean_rows = read_csv(args.clean_data)
    fields = [field for field in fields if field not in EXTRA_FIELDS] + EXTRA_FIELDS
    rows = build_rows(clean_rows, load_workbook_rows(args.workbook), load_lookup_results(args.results_dir))
    stats = dataset_stats(rows)
    write_csv(args.output, fields, rows)
    if args.stats_output:
        args.stats_output.parent.mkdir(parents=True, exist_ok=True)
        args.stats_output.write_text(json.dumps(stats, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(stats, ensure_ascii=False))


if __name__ == "__main__":
    main()
