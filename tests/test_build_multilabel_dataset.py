from __future__ import annotations

import json

import pytest
from openpyxl import Workbook

from scripts.build_multilabel_dataset import build_rows, load_lookup_results, load_workbook_rows


def test_build_rows_preserves_full_hierarchy(tmp_path):
    workbook_path = tmp_path / "patents.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["专利名称", "专利摘要", "授权号"])
    sheet.append(["测试装置", "测试摘要", "CN1A"])
    workbook.save(workbook_path)

    results_dir = tmp_path / "results"
    results_dir.mkdir()
    (results_dir / "cnipa_part_1.json").write_text(
        json.dumps(
            [
                {
                    "row": 2,
                    "status": "ok",
                    "ipc_codes": ["G06F 16/53", "G06N3/08", "G06F16/53"],
                    "found_publication": "CN1A",
                }
            ]
        ),
        encoding="utf-8",
    )

    clean_rows = [
        {
            "patent_title": "测试装置",
            "patent_abstract": "测试摘要",
            "ipc_level_4": "G06F16/53",
        }
    ]
    rows = build_rows(clean_rows, load_workbook_rows(workbook_path), load_lookup_results(results_dir))

    assert rows[0]["publication_numbers"] == "CN1A"
    assert rows[0]["ipc_codes"] == "G06F16/53|G06N3/08"
    assert rows[0]["ipc_sections"] == "G"
    assert rows[0]["ipc_classes"] == "G06"
    assert rows[0]["ipc_subclasses"] == "G06F|G06N"
    assert rows[0]["ipc_code_count"] == "2"


def test_lookup_retries_cannot_disagree(tmp_path):
    results_dir = tmp_path / "results"
    results_dir.mkdir()
    (results_dir / "cnipa_part_1.json").write_text(
        json.dumps(
            [
                {"row": 2, "status": "ok", "ipc_codes": ["G06F16/53"]},
                {"row": 2, "status": "ok", "ipc_codes": ["G06N3/08"]},
            ]
        ),
        encoding="utf-8",
    )

    with pytest.raises(ValueError, match="Conflicting successful IPC results"):
        load_lookup_results(results_dir)


def test_duplicate_source_texts_cannot_disagree():
    clean_rows = [
        {
            "patent_title": "测试装置",
            "patent_abstract": "测试摘要",
            "ipc_level_4": "G06F16/53",
        }
    ]
    workbook_rows = {
        ("测试装置", "测试摘要"): [
            {"row": 2, "publication": "CN1A"},
            {"row": 3, "publication": "CN1B"},
        ]
    }
    lookup_results = {
        2: {"ipc_codes": ["G06F16/53"]},
        3: {"ipc_codes": ["G06F16/53", "G06N3/08"]},
    }

    with pytest.raises(ValueError, match="Duplicate patent rows disagree"):
        build_rows(clean_rows, workbook_rows, lookup_results)
